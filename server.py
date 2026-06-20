from __future__ import annotations

import json
import os
import re
from copy import copy
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parent
TEMPLATE = ROOT / "phd-proposal-review-rubric.xlsx"


def safe_filename(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "review"


def build_comments(payload: dict) -> str:
    feedback = payload.get("feedback", {})
    evidence = payload.get("evidence", [])
    lines = [
        f"Proposal: {payload.get('proposalTitle', '')}",
        f"Field: {payload.get('field', '')}",
        f"Recommendation: {payload.get('recommendation', '')}",
        "",
        f"Strength: {feedback.get('strength', '')}",
        f"Concern: {feedback.get('concern', '')}",
        f"Required revision: {feedback.get('revision', '')}",
        f"Final recommendation: {feedback.get('finalRecommendation', '')}",
        "",
        "Evidence / reviewer notes:",
    ]

    for item in evidence:
        criterion = item.get("criterion", "")
        score = item.get("score", "")
        note = item.get("evidence", "")
        lines.append(f"{criterion}: {score}/5" + (f" - {note}" if note else ""))

    return "\n".join(lines)


def create_review_workbook(payload: dict) -> bytes:
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    scores = payload.get("scores", [])
    for index, value in enumerate(scores[:5], start=3):
        ws[f"I{index}"] = float(value)

    if "H12:I13" in [str(item) for item in ws.merged_cells.ranges]:
        ws["H12"] = payload.get("recommendation", "")
        ws["H12"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws["H12"].font = copy(ws["D12"].font)

    ws["D14"] = build_comments(payload)
    ws["D14"].alignment = Alignment(vertical="top", wrap_text=True)
    ws["D14"].font = Font(name="Aptos", size=10)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


class Handler(SimpleHTTPRequestHandler):
    def translate_path(self, path: str) -> str:
        path = unquote(path.split("?", 1)[0].split("#", 1)[0])
        path = path.lstrip("/")
        return str(ROOT / path)

    def do_POST(self) -> None:
        if self.path != "/export-review":
            self.send_error(404)
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            content = create_review_workbook(payload)
        except Exception as error:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(error)}).encode("utf-8"))
            return

        filename = f"phd-proposal-review-{safe_filename(payload.get('proposalTitle', 'review'))}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)


def main() -> None:
    host = "0.0.0.0"
    port = int(os.environ.get("PORT", "8010"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"Serving HDR simulator on http://{host}:{port}/")
    server.serve_forever()


if __name__ == "__main__":
    main()
